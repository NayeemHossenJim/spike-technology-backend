from __future__ import annotations

import csv
import gzip
import json
import math
import re
import unicodedata
from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO, StringIO
from uuid import UUID
from zipfile import BadZipFile, ZipFile

import xlrd
from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.xml import DEFUSEDXML

REPORT_ARTIFACT_FORMAT_VERSION = 1
REPORT_MAX_SHEETS = 20
REPORT_MAX_ROWS_PER_SHEET = 100_000
REPORT_MAX_COLUMNS_PER_SHEET = 256
REPORT_MAX_TOTAL_CELLS = 2_000_000
REPORT_MAX_LEADING_EMPTY_ROWS = 100
REPORT_MAX_SCANNED_ROWS_PER_SHEET = REPORT_MAX_ROWS_PER_SHEET + REPORT_MAX_LEADING_EMPTY_ROWS + 1
REPORT_MAX_CELL_CHARACTERS = 32_767
REPORT_MAX_HEADER_CHARACTERS = 255
REPORT_MAX_XLSX_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
REPORT_MAX_XLSX_SHARED_STRINGS_BYTES = 32 * 1024 * 1024
REPORT_MAX_XLSX_STYLES_BYTES = 8 * 1024 * 1024
REPORT_MAX_NORMALIZED_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
REPORT_MAX_NORMALIZED_COMPRESSED_BYTES = 50 * 1024 * 1024
REPORT_MAX_PROFILE_BYTES = 2 * 1024 * 1024

_SAFE_ERROR_CODE = re.compile(r"[a-z][a-z0-9_]{0,63}\Z")
_VALUE_TYPE_ORDER = ("null", "boolean", "integer", "number", "date", "datetime", "time", "string")


class ReportParsingError(Exception):
    """A terminal, safely classified problem in an untrusted report."""

    def __init__(self, code: str) -> None:
        if not _SAFE_ERROR_CODE.fullmatch(code):
            raise ValueError("Report parser error codes must be safe identifiers.")
        super().__init__()
        self.code = code


@dataclass(frozen=True, slots=True)
class ReportParsingContext:
    job_id: UUID
    business_id: UUID
    report_upload_id: UUID
    original_filename: str
    file_extension: str
    content_type: str


@dataclass(frozen=True, slots=True)
class ParsedReportArtifacts:
    normalized_content: bytes
    normalized_sha256: str
    profile_content: bytes
    profile_sha256: str
    record_count: int
    sheet_count: int


@dataclass(slots=True)
class _ColumnProfile:
    index: int
    name: str
    source_name: str | None
    generated: bool
    type_counts: Counter[str] = field(default_factory=Counter)

    def observe(self, value_type: str) -> None:
        self.type_counts[value_type] += 1

    def render(self) -> dict[str, object]:
        null_count = self.type_counts["null"]
        total_count = sum(self.type_counts.values())
        return {
            "index": self.index,
            "name": self.name,
            "source_name": self.source_name,
            "generated": self.generated,
            "null_count": null_count,
            "non_null_count": total_count - null_count,
            "type_counts": {
                value_type: self.type_counts[value_type]
                for value_type in _VALUE_TYPE_ORDER
                if self.type_counts[value_type]
            },
        }


class _BoundedJsonlWriter:
    def __init__(self) -> None:
        self._buffer = BytesIO()
        self._gzip = gzip.GzipFile(
            fileobj=self._buffer,
            mode="wb",
            compresslevel=6,
            mtime=0,
        )
        self._uncompressed_bytes = 0
        self._finished = False

    def write(self, payload: dict[str, object]) -> None:
        try:
            encoded = (
                json.dumps(
                    payload,
                    ensure_ascii=False,
                    allow_nan=False,
                    separators=(",", ":"),
                ).encode("utf-8")
                + b"\n"
            )
        except (TypeError, ValueError, UnicodeEncodeError) as exc:
            raise ReportParsingError("cell_value_invalid") from exc

        next_size = self._uncompressed_bytes + len(encoded)
        if next_size > REPORT_MAX_NORMALIZED_UNCOMPRESSED_BYTES:
            raise ReportParsingError("normalized_output_limit_exceeded")
        self._gzip.write(encoded)
        self._uncompressed_bytes = next_size
        if self._buffer.tell() > REPORT_MAX_NORMALIZED_COMPRESSED_BYTES:
            raise ReportParsingError("normalized_output_limit_exceeded")

    def finish(self) -> bytes:
        if not self._finished:
            self._gzip.close()
            self._finished = True
        content = self._buffer.getvalue()
        if len(content) > REPORT_MAX_NORMALIZED_COMPRESSED_BYTES:
            raise ReportParsingError("normalized_output_limit_exceeded")
        return content

    def close(self) -> None:
        if not self._finished:
            self._gzip.close()
            self._finished = True


def _normalize_value(value: object) -> tuple[object, str]:
    if value is None:
        return None, "null"
    if isinstance(value, bool):
        return value, "boolean"
    if isinstance(value, int):
        return value, "integer"
    if isinstance(value, Decimal):
        try:
            if not value.is_finite():
                raise ReportParsingError("non_finite_number")
            if value == value.to_integral_value():
                return int(value), "integer"
            normalized = float(value)
        except (InvalidOperation, OverflowError, ValueError) as exc:
            raise ReportParsingError("cell_value_invalid") from exc
        if not math.isfinite(normalized):
            raise ReportParsingError("non_finite_number")
        return normalized, "number"
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ReportParsingError("non_finite_number")
        if value.is_integer():
            return int(value), "integer"
        return value, "number"
    if isinstance(value, datetime):
        return value.isoformat(), "datetime"
    if isinstance(value, date):
        return value.isoformat(), "date"
    if isinstance(value, time):
        return value.isoformat(), "time"
    if isinstance(value, bytes):
        try:
            value = value.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise ReportParsingError("cell_text_invalid") from exc
    elif not isinstance(value, str):
        value = str(value)

    if not value.strip():
        return None, "null"
    if len(value) > REPORT_MAX_CELL_CHARACTERS:
        raise ReportParsingError("cell_text_limit_exceeded")
    try:
        value.encode("utf-8")
    except UnicodeEncodeError as exc:
        raise ReportParsingError("cell_text_invalid") from exc
    return value, "string"


def _header_text(value: object, *, first_column: bool) -> str | None:
    normalized, _ = _normalize_value(value)
    if normalized is None:
        return None
    text = str(normalized)
    if first_column:
        text = text.removeprefix("\ufeff")
    text = unicodedata.normalize("NFC", " ".join(text.split())).strip()
    if not text:
        return None
    if len(text) > REPORT_MAX_HEADER_CHARACTERS:
        raise ReportParsingError("header_text_limit_exceeded")
    if any(unicodedata.category(character) in {"Cc", "Cf", "Cs"} for character in text):
        raise ReportParsingError("header_text_invalid")
    return text


def _build_columns(header_values: Sequence[object]) -> list[_ColumnProfile]:
    columns: list[_ColumnProfile] = []
    used_names: set[str] = set()
    for index, value in enumerate(header_values):
        source_name = _header_text(value, first_column=index == 0)
        generated = source_name is None
        base_name = source_name or f"column_{index + 1}"
        candidate = base_name
        suffix = 2
        while candidate.casefold() in used_names:
            candidate = f"{base_name}__{suffix}"
            suffix += 1
        used_names.add(candidate.casefold())
        columns.append(
            _ColumnProfile(
                index=index,
                name=candidate,
                source_name=source_name,
                generated=generated,
            )
        )
    return columns


def _row_is_empty(values: Sequence[object]) -> bool:
    return all(_normalize_value(value)[0] is None for value in values)


class _ReportNormalizer:
    def __init__(self) -> None:
        self.writer = _BoundedJsonlWriter()
        self.sheets: list[dict[str, object]] = []
        self.record_count = 0
        self.total_cells = 0

    def add_sheet(
        self,
        *,
        source_sheet_index: int,
        sheet_name: str,
        rows: Iterable[tuple[int, Sequence[object]]],
    ) -> None:
        if len(self.sheets) >= REPORT_MAX_SHEETS:
            raise ReportParsingError("sheet_limit_exceeded")

        header_row_number: int | None = None
        columns: list[_ColumnProfile] | None = None
        sheet_record_count = 0
        scanned_rows = 0

        for row_number, raw_values in rows:
            scanned_rows += 1
            if scanned_rows > REPORT_MAX_SCANNED_ROWS_PER_SHEET:
                raise ReportParsingError("row_limit_exceeded")
            if len(raw_values) > REPORT_MAX_COLUMNS_PER_SHEET:
                raise ReportParsingError("column_limit_exceeded")

            if columns is None:
                if _row_is_empty(raw_values):
                    if scanned_rows > REPORT_MAX_LEADING_EMPTY_ROWS:
                        raise ReportParsingError("header_not_found")
                    continue
                trimmed_header = list(raw_values)
                while trimmed_header and _normalize_value(trimmed_header[-1])[0] is None:
                    trimmed_header.pop()
                if not trimmed_header:
                    continue
                columns = _build_columns(trimmed_header)
                header_row_number = row_number
                continue

            normalized_values = [_normalize_value(value) for value in raw_values]
            extra_values = normalized_values[len(columns) :]
            if any(value is not None for value, _ in extra_values):
                raise ReportParsingError("row_width_exceeded")
            normalized_values = normalized_values[: len(columns)]
            normalized_values.extend([(None, "null")] * (len(columns) - len(normalized_values)))
            if all(value is None for value, _ in normalized_values):
                continue
            if sheet_record_count >= REPORT_MAX_ROWS_PER_SHEET:
                raise ReportParsingError("row_limit_exceeded")

            next_total_cells = self.total_cells + len(columns)
            if next_total_cells > REPORT_MAX_TOTAL_CELLS:
                raise ReportParsingError("cell_limit_exceeded")

            values: dict[str, object] = {}
            for column, (value, value_type) in zip(columns, normalized_values, strict=True):
                column.observe(value_type)
                values[column.name] = value

            self.writer.write(
                {
                    "sheet_index": source_sheet_index,
                    "sheet_name": sheet_name,
                    "row_number": row_number,
                    "values": values,
                }
            )
            sheet_record_count += 1
            self.record_count += 1
            self.total_cells = next_total_cells

        if columns is None or header_row_number is None:
            return
        self.sheets.append(
            {
                "index": source_sheet_index,
                "name": sheet_name,
                "header_row_number": header_row_number,
                "record_count": sheet_record_count,
                "columns": [column.render() for column in columns],
            }
        )

    def finish(self) -> tuple[bytes, list[dict[str, object]], int]:
        if not self.sheets:
            raise ReportParsingError("workbook_has_no_data_sheets")
        return self.writer.finish(), self.sheets, self.record_count

    def close(self) -> None:
        self.writer.close()


def _parse_csv(content: bytes, normalizer: _ReportNormalizer) -> None:
    try:
        decoded = content.decode("utf-8-sig")
        reader = csv.reader(StringIO(decoded, newline=""), strict=True)
        normalizer.add_sheet(
            source_sheet_index=0,
            sheet_name="data",
            rows=((row_number, row) for row_number, row in enumerate(reader, start=1)),
        )
    except ReportParsingError:
        raise
    except (csv.Error, UnicodeDecodeError) as exc:
        raise ReportParsingError("csv_malformed") from exc


def _xlsx_preflight(content: bytes) -> None:
    if not DEFUSEDXML:
        raise RuntimeError("openpyxl must use defusedxml for untrusted workbooks")
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if sum(entry.file_size for entry in entries) > (REPORT_MAX_XLSX_UNCOMPRESSED_BYTES):
                raise ReportParsingError("workbook_expansion_limit_exceeded")
            entry_sizes = {entry.filename.casefold(): entry.file_size for entry in entries}
            if entry_sizes.get("xl/sharedstrings.xml", 0) > (REPORT_MAX_XLSX_SHARED_STRINGS_BYTES):
                raise ReportParsingError("workbook_metadata_limit_exceeded")
            if entry_sizes.get("xl/styles.xml", 0) > REPORT_MAX_XLSX_STYLES_BYTES:
                raise ReportParsingError("workbook_metadata_limit_exceeded")
    except ReportParsingError:
        raise
    except (BadZipFile, OSError) as exc:
        raise ReportParsingError("xlsx_malformed") from exc


def _parse_xlsx(content: bytes, normalizer: _ReportNormalizer) -> None:
    _xlsx_preflight(content)
    workbook = None
    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
            rich_text=False,
        )
        worksheets = workbook.worksheets
        if len(worksheets) > REPORT_MAX_SHEETS:
            raise ReportParsingError("sheet_limit_exceeded")
        for sheet_index, worksheet in enumerate(worksheets):
            if (worksheet.max_column or 0) > REPORT_MAX_COLUMNS_PER_SHEET:
                raise ReportParsingError("column_limit_exceeded")
            if (worksheet.max_row or 0) > REPORT_MAX_SCANNED_ROWS_PER_SHEET:
                raise ReportParsingError("row_limit_exceeded")
            normalizer.add_sheet(
                source_sheet_index=sheet_index,
                sheet_name=worksheet.title,
                rows=(
                    (row_number, tuple(cell.value for cell in row))
                    for row_number, row in enumerate(worksheet.iter_rows(), start=1)
                ),
            )
    except ReportParsingError:
        raise
    except (
        BadZipFile,
        DefusedXmlException,
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
    ) as exc:
        raise ReportParsingError("xlsx_malformed") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _xls_cell_value(cell: xlrd.sheet.Cell, *, datemode: int) -> object:
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        except (OverflowError, TypeError, ValueError, xlrd.XLDateError) as exc:
            raise ReportParsingError("cell_value_invalid") from exc
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code.get(cell.value, "#ERROR")
    return cell.value


def _parse_xls(content: bytes, normalizer: _ReportNormalizer) -> None:
    workbook = None
    try:
        workbook = xlrd.open_workbook(
            file_contents=content,
            on_demand=True,
            ragged_rows=True,
            formatting_info=False,
            logfile=StringIO(),
            verbosity=0,
        )
        if workbook.nsheets > REPORT_MAX_SHEETS:
            raise ReportParsingError("sheet_limit_exceeded")
        for sheet_index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(sheet_index)
            if worksheet.ncols > REPORT_MAX_COLUMNS_PER_SHEET:
                raise ReportParsingError("column_limit_exceeded")
            if worksheet.nrows > REPORT_MAX_SCANNED_ROWS_PER_SHEET:
                raise ReportParsingError("row_limit_exceeded")
            normalizer.add_sheet(
                source_sheet_index=sheet_index,
                sheet_name=worksheet.name,
                rows=(
                    (
                        row_number + 1,
                        tuple(
                            _xls_cell_value(cell, datemode=workbook.datemode)
                            for cell in worksheet.row(row_number)
                        ),
                    )
                    for row_number in range(worksheet.nrows)
                ),
            )
    except ReportParsingError:
        raise
    except (EOFError, OSError, TypeError, ValueError, xlrd.XLRDError) as exc:
        raise ReportParsingError("xls_malformed") from exc
    finally:
        if workbook is not None:
            workbook.release_resources()


def parse_report_content(
    *,
    context: ReportParsingContext,
    content: bytes,
) -> ParsedReportArtifacts:
    """Parse one immutable upload into deterministic, bounded private artifacts."""

    normalizer = _ReportNormalizer()
    try:
        if context.file_extension == "csv":
            _parse_csv(content, normalizer)
        elif context.file_extension == "xlsx":
            _parse_xlsx(content, normalizer)
        elif context.file_extension == "xls":
            _parse_xls(content, normalizer)
        else:
            raise ReportParsingError("unsupported_report_format")
        normalized_content, sheets, record_count = normalizer.finish()
    except Exception:
        normalizer.close()
        raise

    normalized_digest = sha256(normalized_content).hexdigest()
    profile = {
        "format": "spike.report-profile",
        "version": REPORT_ARTIFACT_FORMAT_VERSION,
        "job_id": str(context.job_id),
        "business_id": str(context.business_id),
        "source": {
            "report_upload_id": str(context.report_upload_id),
            "filename": context.original_filename,
            "extension": context.file_extension,
            "content_type": context.content_type,
        },
        "normalized": {
            "format": "jsonl",
            "compression": "gzip",
            "sha256": normalized_digest,
        },
        "record_count": record_count,
        "sheet_count": len(sheets),
        "sheets": sheets,
    }
    try:
        profile_content = (
            json.dumps(
                profile,
                ensure_ascii=False,
                allow_nan=False,
                separators=(",", ":"),
                sort_keys=True,
            ).encode("utf-8")
            + b"\n"
        )
    except (TypeError, ValueError, UnicodeEncodeError) as exc:
        raise ReportParsingError("profile_output_invalid") from exc
    if len(profile_content) > REPORT_MAX_PROFILE_BYTES:
        raise ReportParsingError("profile_output_limit_exceeded")

    return ParsedReportArtifacts(
        normalized_content=normalized_content,
        normalized_sha256=normalized_digest,
        profile_content=profile_content,
        profile_sha256=sha256(profile_content).hexdigest(),
        record_count=record_count,
        sheet_count=len(sheets),
    )


__all__ = [
    "REPORT_ARTIFACT_FORMAT_VERSION",
    "REPORT_MAX_CELL_CHARACTERS",
    "REPORT_MAX_COLUMNS_PER_SHEET",
    "REPORT_MAX_ROWS_PER_SHEET",
    "REPORT_MAX_SHEETS",
    "REPORT_MAX_TOTAL_CELLS",
    "ParsedReportArtifacts",
    "ReportParsingContext",
    "ReportParsingError",
    "parse_report_content",
]
